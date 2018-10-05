#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import print_function

import logging
import traceback
import csv
from cStringIO import StringIO
from openpyxl import load_workbook
from rdflib import Graph
from rdflib.namespace import SKOS, RDF

from ckan.common import _, ungettext
import ckan.plugins.toolkit as toolkit

from ckan.lib.base import config
from ckan.lib.cli import CkanCommand

from ckanext.faociok.models import (Vocabulary, VocabularyTerm,
                                    Session, load_vocabulary,
                                    find_unused_terms, setup_models)

log = logging.getLogger(__name__)


class VocabularyCommands(CkanCommand):
    """Manage vocabularies in FAO-CIOK extension.
    """

    summary = __doc__.split('\n')[0]
    usage = __doc__

    @property
    def usage(self):
        out = [self.__doc__]
        commands = self.get_commands()
        for ckey in sorted(commands.keys()):
            cmd = commands[ckey]
            cusage = self.command_usage(ckey, cmd)
            out.append(cusage)
        return '\n'.join(out)

    def cmd_initdb(self, *args, **kwargs):
        setup_models()

    def cmd_list(self, *args, **kwargs):
        """
        List vocabularies
        """
        print(_('Vocabularies:'))
        for voc in Vocabulary.get_all():
            print(_('vocabulary name: {}').format(voc.name))
            print(_('  has relations: {}').format(voc.has_relations))
            print()
        print(_('[end of vocabularies list]'))

    def cmd_create(self, vocabulary_name, has_relations=False, *args, **kwargs):
        """
        Create vocabulary

        syntax: create vocabulary_name [has_relations,default=no]

        """
        Vocabulary.create(vocabulary_name, bool(has_relations))
         
    def cmd_load(self, vocabulary_name, path, *args, **kwargs):
        """
        Load vocabulary data

        syntax: load vocabulary_name path_to_source_file
        """
        with open(path, 'rt') as f:
            count = load_vocabulary(vocabulary_name, f)
            print(_('loaded {} terms from {} to {} vocabulary').format(count, path, vocabulary_name))

    def cmd_import_agrovoc(self, in_file, *args, **kwargs):
        """
        Import AGROVOC terms from RDF file

        syntax: import_agrovoc rdf_file
        """
        OFFERED_LANGS = (config.get('ckan.locales_offered') or 'en es fr de it').lower().split(' ')
        
        header = ('parent', 'term',) + tuple(('lang:{}'.format(L) for L in OFFERED_LANGS)) + ( 'property:parents',)
        rdata = []
        rdata.append(header)

        g = Graph()
        g.parse(in_file, format='nt')
        for o, p, s in g.triples((None, RDF.type, SKOS.Concept)):
            cid = str(o).split('/')[-1]
            row = {'term': cid}

            for label_r in g.triples((o, SKOS.prefLabel, None)):
                label = label_r[2]
                if not label.language in OFFERED_LANGS:
                    continue
                row['lang:{}'.format(label.language)] = label.value

            iparents = g.triples((o, SKOS.broader, None))
            parents = []
            for to in iparents:
                parent = to[-1]
                row['parent'] = str(parent).split('/')[-1]
                parents.append(row['parent'])
            if not parents:
                row['parent'] = None
            row['property:parents'] = ','.join(parents)

            row_data = []
            for col in header:
                if col.startswith('lang'):
                    val = row.get(col) or row.get('lang:en') or row.get('lang:fr') 
                else:
                    val = row[col]
                row_data.append(val.encode('utf-8') if isinstance(val, unicode) else val)
            
            if row['parent']:
                rdata.append(row_data)
            else:
                # top-level should be first
                rdata.insert(1, row_data)

        log.info('AGROVOC terms parsed: %s', len(rdata))
        csvdata = StringIO()
        w = csv.writer(csvdata)
        w.writerows(rdata)

        csvdata.seek(0)
        voc_name = Vocabulary.VOCABULARY_AGROVOC
        try:
            voc = Vocabulary.get(voc_name)
        except ValueError:
            voc = Vocabulary.create(voc_name, has_relations=True)

        count = load_vocabulary(voc_name, csvdata)
        log.info('AGROVOC terms imported: %s', count)
        cleanup_stats = find_unused_terms(voc_name, 'fao_agrovoc')
        if cleanup_stats['datasets']:
            print("Following dataset have terms not present in vocabulary:")
            for dname, tvals in sorted(cleanup_stats['datasets'].items()):
                print(' dataset', dname,':', ','.join(tvals))

    def cmd_import_m49(self, in_file, *args, **kwargs):
        """
        Convert xlsx file with m49 data into vocabulary

        syntax: import_m49 in_file
        """
        wb = load_workbook(in_file)
        sheet = wb.active

        IDX_COUNTRY_M49 = 1
        IDX_COUNTRY_ISO3 = 2
        IDX_COUNTRY_NAME = 3
        IDX_L1_M49 = 8
        IDX_L1_NAME = 9
        #IDX_L2_CODE = 10
        #IDX_L2_NAME = 11

        level1_cells = (IDX_L1_M49, IDX_L1_NAME,)
        countries_cells = (IDX_COUNTRY_M49, IDX_COUNTRY_NAME, IDX_COUNTRY_ISO3)
        countries_parent_cell = IDX_L1_M49

        countries = {}  # key: id , value : row
        level1 = {}     # key: id , value : row

        for row in sheet.iter_rows(min_row=6):
            for indexes, parent_idx, container in ((countries_cells, countries_parent_cell, countries,),
                                                   (level1_cells, None, level1,)):
                # ( parent, data..)
                rdata = []
                rdata.append(row[parent_idx-1].value if parent_idx else None)
                
                for idx in indexes:
                    value = row[idx-1].value
                    try:
                        value = value.replace('(M49)', '').replace('(MDG=M49)', '') if value else value
                    except AttributeError:
                        # prolly it has been parsed as a number
                        pass
                    rdata.append(value.encode('utf-8') if isinstance(value, unicode) else value)

                if not any(rdata):
                    continue

                id = rdata[1]
                container[id] = rdata  # l1 and l2 may be repeated

        csvdata = StringIO()
        w = csv.writer(csvdata)
        w.writerow(['parent', 'term', 'property:country_code', 'lang:en', 'lang:fr', 'lang:es'])
        for id,r in level1.iteritems():
            # print('L1 ROW {}: {}'.format(id,r))
            w.writerow([r[0]] + [r[1]] + [""] + [r[2]] + [r[2]+" [FR]"] + [r[2]+" [ES]"] )
        for id,r in countries.iteritems():
            # print('CNTY ROW {}: {}'.format(id,r))
            w.writerow([r[0]] + [r[1]] + [r[3]] + [r[2]] + [r[2]+" [FR]"] + [r[2]+" [ES]"] )
        csvdata.seek(0)
        voc_name = Vocabulary.VOCABULARY_M49_REGIONS

        try:
            voc = Vocabulary.get(voc_name)
        except ValueError:
            voc = Vocabulary.create(voc_name, has_relations=True)

        count = load_vocabulary(voc_name, csvdata)
        print(_('loaded {} terms from {} to {} vocabulary').format(count, in_file, voc_name))


    def get_commands(self):
        """
        Return dictionary of command-> callable 
        """
        return dict([ (n[4:], getattr(self, n),) for n in dir(self) if n.startswith('cmd_')])
        
    def command_usage(self, cmd, callable):
        return '    {}\n     {}\n'.format(cmd, (callable.__doc__ or 'No documentation yet').strip())

    def command(self):
        '''
        Parse command line arguments and call appropriate method.
        '''
        try:
            cmd = self.args[0]
        except IndexError:
            print(_("ERROR: missing command"))
            print(self.usage)
            return

        self._load_config()

        commands = self.get_commands()
        if not cmd in commands:
            print(_("ERROR: invalid command: {}").format(cmd))
            print(self.usage)
            return
        callable = commands[cmd]
        try:
            out = callable(*self.args[1:], **vars(self.options))
            Session.commit()
            return out
        except Exception, err:
            log.error(_("Can't execute %s with args %s: %s"), cmd, self.args[1:], err, exc_info=err)
            traceback.print_exc(err)
            print(_("Can't execute {} with args {}: {}").format(cmd, self.args[1:], err))
            print(self.command_usage(cmd, callable))
            return
